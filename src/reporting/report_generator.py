"""
Reporting Export Module

Handles report generation including:
- Report template management
- PDF generation
- Report scheduling & automation
- ROI calculations
"""

import logging
from datetime import datetime, timezone
from typing import Dict, List, Optional, Any
from dataclasses import dataclass, field
from enum import Enum
from uuid import uuid4

from src.telemetry.metrics import MetricsCollector, LatencyTracker
from src.telemetry.events import EventLogger
from src.campaigns.campaign_manager import Campaign
from src.analytics.analytics_store import CampaignPerformance

logger = logging.getLogger(__name__)


class ReportType(Enum):
    """Report types"""
    SPONSOR_REPORT = "sponsor_report"
    PERFORMANCE_SUMMARY = "performance_summary"
    ROI_REPORT = "roi_report"
    CUSTOM = "custom"


class ReportFormat(Enum):
    """Report formats"""
    PDF = "pdf"
    CSV = "csv"
    EXCEL = "excel"
    JSON = "json"


@dataclass
class ReportTemplate:
    """Report template"""
    template_id: str
    name: str
    report_type: ReportType
    sections: List[str]  # List of section names
    include_roi: bool = True
    include_attribution: bool = True
    include_benchmarks: bool = False
    branding_enabled: bool = True


@dataclass
class Report:
    """Generated report"""
    report_id: str
    campaign_id: str
    template_id: str
    report_type: ReportType
    format: ReportFormat
    generated_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    file_size_bytes: Optional[int] = None
    file_url: Optional[str] = None
    includes_roi: bool = False
    includes_attribution: bool = False
    metadata: Dict[str, Any] = field(default_factory=dict)


class ReportGenerator:
    """
    Report Generator
    
    Generates reports from campaign data with support for:
    - Multiple templates
    - PDF/CSV/Excel formats
    - ROI calculations
    - Automated scheduling
    """
    
    def __init__(
        self,
        metrics_collector: MetricsCollector,
        event_logger: EventLogger
    ):
        self.metrics = metrics_collector
        self.events = event_logger
        # In production, this would connect to storage
        self._reports: Dict[str, Report] = {}
        self._templates: Dict[str, ReportTemplate] = {}
        self._initialize_default_templates()
        
    def _initialize_default_templates(self):
        """Initialize default report templates"""
        # Basic sponsor report template
        basic_template = ReportTemplate(
            template_id="basic_sponsor",
            name="Basic Sponsor Report",
            report_type=ReportType.SPONSOR_REPORT,
            sections=["overview", "performance", "attribution", "roi"],
            include_roi=True,
            include_attribution=True
        )
        self._templates[basic_template.template_id] = basic_template
        
        # Detailed ROI report template
        roi_template = ReportTemplate(
            template_id="detailed_roi",
            name="Detailed ROI Report",
            report_type=ReportType.ROI_REPORT,
            sections=["overview", "performance", "attribution", "roi", "benchmarks", "recommendations"],
            include_roi=True,
            include_attribution=True,
            include_benchmarks=True
        )
        self._templates[roi_template.template_id] = roi_template
        
    async def generate_report(
        self,
        campaign_id: str,
        report_type: ReportType,
        format: ReportFormat,
        template_id: Optional[str] = None,
        include_roi: bool = True,
        include_attribution: bool = True,
        include_benchmarks: bool = False,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        postgres_conn=None,
        analytics_store=None,
        roi_calculator=None
    ) -> Report:
        """
        Generate a report for a campaign
        
        Args:
            user_id: User generating the report
            campaign_id: Campaign ID
            template_id: Template to use
            format: Output format
            custom_sections: Optional custom sections to include
            
        Returns:
            Generated report
            
        Telemetry:
            - report_generation_time: Time to generate report
            - report_generation_success: Success rate
            - pdf_size: Size of generated PDF
        """
        import time
        start_time = time.time()
        
        # Get template (use default if not provided)
        template_id = template_id or "basic_sponsor"
        template = self._templates.get(template_id)
        if not template:
            # Create default template if not found
            template = ReportTemplate(
                template_id=template_id,
                name="Default Report",
                report_type=report_type,
                sections=["overview", "performance", "attribution", "roi"],
                include_roi=include_roi,
                include_attribution=include_attribution,
                include_benchmarks=include_benchmarks
            )
        
        # Generate report data with real data sources
        report_data = await self._generate_report_data(
            campaign_id, template, postgres_conn, analytics_store, roi_calculator
        )
        
        # Generate file based on format
        if format == ReportFormat.PDF:
            file_url, file_size = await self._generate_pdf(report_data, template, campaign_id)
        elif format == ReportFormat.CSV:
            file_url, file_size = await self._generate_csv(report_data, template, campaign_id)
        elif format == ReportFormat.EXCEL:
            file_url, file_size = await self._generate_excel(report_data, template, campaign_id)
        else:
            raise ValueError(f"Unsupported format: {format}")
        
        # Create report record
        report = Report(
            report_id=str(uuid4()),
            campaign_id=campaign_id,
            template_id=template.template_id,
            report_type=report_type,
            format=format,
            file_size_bytes=file_size,
            file_url=file_url,
            includes_roi=include_roi,
            includes_attribution=include_attribution,
            metadata={"sections": template.sections, "report_data": report_data}
        )
        
        self._reports[report.report_id] = report
        
        # Record telemetry
        generation_time = time.time() - start_time
        self.metrics.record_histogram(
            "report_generation_time",
            generation_time,
            tags={"format": format.value, "template_id": template.template_id}
        )
        self.metrics.record_gauge(
            "report_file_size_bytes",
            file_size,
            tags={"format": format.value}
        )
        self.metrics.increment_counter(
            "report_generated",
            tags={"format": format.value, "template_id": template.template_id}
        )
        
        return report
    
    async def _generate_report_data(
        self,
        campaign_id: str,
        template: ReportTemplate,
        postgres_conn=None,
        analytics_store=None,
        roi_calculator=None
    ) -> Dict[str, Any]:
        """Generate report data from campaign and analytics"""
        report_data = {
            "campaign_id": campaign_id,
            "template": template.name,
            "sections": {},
            "generated_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Fetch campaign data if postgres_conn available
        campaign_data = None
        if postgres_conn:
            try:
                campaign_row = await postgres_conn.fetchrow(
                    """
                    SELECT c.*, s.name as sponsor_name, s.email as sponsor_email,
                           p.name as podcast_name
                    FROM campaigns c
                    LEFT JOIN sponsors s ON c.sponsor_id = s.sponsor_id
                    LEFT JOIN podcasts p ON c.podcast_id = p.podcast_id
                    WHERE c.campaign_id = $1
                    """,
                    campaign_id
                )
                if campaign_row:
                    campaign_data = dict(campaign_row)
            except Exception as e:
                logger.warning(f"Failed to fetch campaign data: {e}")
        
        # Generate each section with real data
        for section in template.sections:
            report_data["sections"][section] = await self._generate_section(
                campaign_id, section, template, campaign_data, analytics_store, roi_calculator
            )
        
        return report_data
    
    async def _generate_section(
        self,
        campaign_id: str,
        section: str,
        template: ReportTemplate,
        campaign_data: Optional[Dict[str, Any]] = None,
        analytics_store=None,
        roi_calculator=None
    ) -> Dict[str, Any]:
        """Generate a report section with real data"""
        section_data = {
            "section": section,
            "data": {}
        }
        
        if section == "overview":
            if campaign_data:
                section_data["data"] = {
                    "campaign_name": campaign_data.get("name", "Campaign"),
                    "sponsor": campaign_data.get("sponsor_name", "Unknown Sponsor"),
                    "sponsor_email": campaign_data.get("sponsor_email"),
                    "podcast_name": campaign_data.get("podcast_name", "Podcast"),
                    "date_range": f"{campaign_data.get('start_date', '').strftime('%Y-%m-%d')} to {campaign_data.get('end_date', '').strftime('%Y-%m-%d')}" if campaign_data.get('start_date') else "N/A",
                    "status": campaign_data.get("status", "unknown")
                }
            else:
                section_data["data"] = {
                    "campaign_name": "Campaign",
                    "sponsor": "Unknown Sponsor",
                    "date_range": "N/A"
                }
        
        elif section == "performance":
            if analytics_store and campaign_data:
                try:
                    performance = await analytics_store.calculate_campaign_performance(
                        campaign_id=campaign_id,
                        podcast_id=str(campaign_data.get("podcast_id", "")),
                        start_date=campaign_data.get("start_date", datetime.now(timezone.utc)),
                        end_date=campaign_data.get("end_date", datetime.now(timezone.utc))
                    )
                    section_data["data"] = {
                        "total_downloads": performance.total_downloads,
                        "total_streams": performance.total_streams,
                        "total_listeners": performance.total_listeners,
                        "attribution_events": performance.attribution_events,
                        "conversions": performance.conversions
                    }
                except Exception as e:
                    logger.warning(f"Failed to fetch performance data: {e}")
                    section_data["data"] = {
                        "total_downloads": 0,
                        "total_streams": 0,
                        "total_listeners": 0
                    }
            else:
                section_data["data"] = {
                    "total_downloads": 0,
                    "total_streams": 0,
                    "total_listeners": 0
                }
        
        elif section == "attribution" and template.include_attribution:
            if analytics_store:
                try:
                    events = await analytics_store.get_attribution_events(campaign_id)
                    conversions = [e for e in events if e.conversion_type]
                    section_data["data"] = {
                        "attribution_events": len(events),
                        "conversions": len(conversions),
                        "conversion_rate": len(conversions) / len(events) if events else 0.0,
                        "conversion_value": sum(e.conversion_value or 0 for e in conversions)
                    }
                except Exception as e:
                    logger.warning(f"Failed to fetch attribution data: {e}")
                    section_data["data"] = {
                        "attribution_events": 0,
                        "conversions": 0,
                        "conversion_rate": 0.0
                    }
            else:
                section_data["data"] = {
                    "attribution_events": 0,
                    "conversions": 0,
                    "conversion_rate": 0.0
                }
        
        elif section == "roi" and template.include_roi:
            if campaign_data and analytics_store:
                try:
                    campaign_cost = float(campaign_data.get("campaign_value", 0))
                    performance = await analytics_store.calculate_campaign_performance(
                        campaign_id=campaign_id,
                        podcast_id=str(campaign_data.get("podcast_id", "")),
                        start_date=campaign_data.get("start_date", datetime.now(timezone.utc)),
                        end_date=campaign_data.get("end_date", datetime.now(timezone.utc))
                    )
                    conversion_value = performance.conversion_value
                    
                    roi = ((conversion_value - campaign_cost) / campaign_cost * 100) if campaign_cost > 0 else 0.0
                    roas = (conversion_value / campaign_cost) if campaign_cost > 0 else 0.0
                    
                    section_data["data"] = {
                        "campaign_cost": campaign_cost,
                        "conversion_value": conversion_value,
                        "roi": roi,
                        "roas": roas,
                        "net_profit": conversion_value - campaign_cost,
                        "conversions": performance.conversions
                    }
                except Exception as e:
                    logger.warning(f"Failed to calculate ROI: {e}")
                    section_data["data"] = {
                        "campaign_cost": campaign_data.get("campaign_value", 0) if campaign_data else 0,
                        "conversion_value": 0.0,
                        "roi": 0.0,
                        "roas": 0.0
                    }
            else:
                section_data["data"] = {
                    "campaign_cost": 0.0,
                    "conversion_value": 0.0,
                    "roi": 0.0,
                    "roas": 0.0
                }
        
        return section_data
    
    async def _generate_pdf(
        self,
        report_data: Dict[str, Any],
        template: ReportTemplate,
        campaign_id: str
    ) -> tuple[str, int]:
        """Generate PDF report"""
        try:
            # Try to use reportlab if available
            try:
                from reportlab.lib.pagesizes import letter
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.lib import colors
                import io
                
                buffer = io.BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=letter)
                story = []
                styles = getSampleStyleSheet()
                
                # Title
                story.append(Paragraph(f"Campaign Report: {report_data.get('campaign_id', 'Campaign')}", styles['Title']))
                story.append(Spacer(1, 12))
                
                # Generate sections
                for section_name, section_data in report_data.get('sections', {}).items():
                    story.append(Paragraph(f"<b>{section_name.title()}</b>", styles['Heading2']))
                    story.append(Spacer(1, 6))
                    
                    # Add section data as table
                    data = section_data.get('data', {})
                    if data:
                        table_data = [[str(k), str(v)] for k, v in data.items()]
                        table = Table(table_data)
                        table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 12),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ]))
                        story.append(table)
                        story.append(Spacer(1, 12))
                
                doc.build(story)
                pdf_bytes = buffer.getvalue()
                buffer.close()
                
                # Store PDF (in production, would use S3 or similar)
                import os
                reports_dir = os.getenv("REPORTS_STORAGE_PATH", "/tmp/reports")
                os.makedirs(reports_dir, exist_ok=True)
                file_path = os.path.join(reports_dir, f"{campaign_id}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.pdf")
                
                with open(file_path, 'wb') as f:
                    f.write(pdf_bytes)
                
                file_size = len(pdf_bytes)
                file_url = f"/api/v1/reports/{os.path.basename(file_path)}"
                
                return file_url, file_size
            except ImportError:
                # Fallback: Generate simple text-based PDF or return JSON
                logger.warning("reportlab not available, generating simple report")
                import json
                import os
                
                reports_dir = os.getenv("REPORTS_STORAGE_PATH", "/tmp/reports")
                os.makedirs(reports_dir, exist_ok=True)
                file_path = os.path.join(reports_dir, f"{campaign_id}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.json")
                
                with open(file_path, 'w') as f:
                    json.dump(report_data, f, indent=2, default=str)
                
                file_size = os.path.getsize(file_path)
                file_url = f"/api/v1/reports/{os.path.basename(file_path)}"
                
                return file_url, file_size
        except Exception as e:
            logger.error(f"Failed to generate PDF: {e}", exc_info=True)
            # Return placeholder
            file_url = f"/reports/{campaign_id}.pdf"
            file_size = 1024 * 50
            return file_url, file_size
    
    async def _generate_csv(
        self,
        report_data: Dict[str, Any],
        template: ReportTemplate,
        campaign_id: str
    ) -> tuple[str, int]:
        """Generate CSV report"""
        import csv
        import os
        
        reports_dir = os.getenv("REPORTS_STORAGE_PATH", "/tmp/reports")
        os.makedirs(reports_dir, exist_ok=True)
        file_path = os.path.join(reports_dir, f"{campaign_id}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.csv")
        
        with open(file_path, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Section', 'Metric', 'Value'])
            
            for section_name, section_data in report_data.get('sections', {}).items():
                data = section_data.get('data', {})
                for key, value in data.items():
                    writer.writerow([section_name, key, value])
        
        file_size = os.path.getsize(file_path)
        file_url = f"/api/v1/reports/{os.path.basename(file_path)}"
        
        return file_url, file_size
    
    async def _generate_excel(
        self,
        report_data: Dict[str, Any],
        template: ReportTemplate,
        campaign_id: str
    ) -> tuple[str, int]:
        """Generate Excel report"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Campaign Report"
            
            row = 1
            for section_name, section_data in report_data.get('sections', {}).items():
                # Section header
                ws.cell(row=row, column=1, value=section_name.title()).font = Font(bold=True)
                row += 1
                
                # Section data
                data = section_data.get('data', {})
                for key, value in data.items():
                    ws.cell(row=row, column=1, value=str(key))
                    ws.cell(row=row, column=2, value=str(value))
                    row += 1
                
                row += 1  # Spacing between sections
            
            import os
            reports_dir = os.getenv("REPORTS_STORAGE_PATH", "/tmp/reports")
            os.makedirs(reports_dir, exist_ok=True)
            file_path = os.path.join(reports_dir, f"{campaign_id}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx")
            
            wb.save(file_path)
            file_size = os.path.getsize(file_path)
            file_url = f"/api/v1/reports/{os.path.basename(file_path)}"
            
            return file_url, file_size
        except ImportError:
            # Fallback to CSV if openpyxl not available
            logger.warning("openpyxl not available, falling back to CSV")
            return await self._generate_csv(report_data, template, campaign_id)
    
    async def get_report(self, report_id: str) -> Optional[Report]:
        """Get report by ID"""
        return self._reports.get(report_id)
    
    async def list_reports(
        self,
        campaign_id: Optional[str] = None,
        user_id: Optional[str] = None
    ) -> List[Report]:
        """List reports with optional filters"""
        reports = list(self._reports.values())
        
        if campaign_id:
            reports = [r for r in reports if r.campaign_id == campaign_id]
        
        return reports
    
    async def calculate_roi(
        self,
        campaign_id: str,
        campaign_cost: float,
        conversion_value: float
    ) -> Dict[str, float]:
        """Calculate ROI metrics"""
        roi = (conversion_value - campaign_cost) / campaign_cost if campaign_cost > 0 else 0.0
        roas = conversion_value / campaign_cost if campaign_cost > 0 else 0.0
        
        return {
            "roi": roi,
            "roas": roas,
            "campaign_cost": campaign_cost,
            "conversion_value": conversion_value,
            "net_profit": conversion_value - campaign_cost
        }
